from elasticsearch import Elasticsearch
import json
import time
import re

import openpyxl
from openpyxl.cell import cell
from openpyxl.styles import Alignment




# excel_file_path = 'data/tests_results.xlsx'

es = Elasticsearch("http://os-2828.homecredit.ru:9200")



# Call an API, in this example `info()` /api/datasources/proxy/87/_msearch
# resp = client.info()

ws_methods_control_list = (
'CustomerWSEndpoint#',
'CustomerUpdateWSEndpoint#',
'CustomerUpdWSEndpoint#',
'ClientWSEndpoint#',
'EventWSEndpoint#',
'ScanServerStorageWSEndpoint#'
)



"""
"gte":1642065540000,"lte":1642069140000,
"""

wl_stat_dict = {
    'Total_Client' : {'count': 0, },
    'Total_Customer' : {'count': 0, },
    'Total_Event' : {'count': 0, },
    'Total_Scan' : {'count': 0, }
}

def wl_stat_data_parcer(es, ws_methods_control_list, struct_left_time_bound, struct_right_time_bound, index):

    for method in ws_methods_control_list:

        query_pers_getcutomerdata = {"size":0,"query":{"bool":{"filter":[{"range":{"@timestamp":{"gte":struct_left_time_bound,"lte":struct_right_time_bound,"format":"epoch_millis"}}},{"query_string":{"analyze_wildcard":'true',"query": method}}]}},"aggs":{"2":{"terms":{"field":"request.keyword","size":500,"order":{"_term":"desc"},"min_doc_count":1},"aggs":{"3":{"date_histogram":{"interval":"1m","field":"@timestamp","min_doc_count":"1","extended_bounds":{"min":1642065540000,"max":1642069140000},"format":"epoch_millis"},"aggs":{"1":{"percentiles":{"field":"duration","percents":["90"],"script":"_value*1000"}}}}}}}}

        res = es.search(index=index, body=query_pers_getcutomerdata)

        for p in res['aggregations']['2']['buckets']:
            method_name = p['key']
            method_count = p['doc_count']

            # подсчет количества исполнений методов (Total_*)
            ws_name_from_method_name = re.sub(r'([A-Z])', r' \1', method_name).split()[0]
            wl_stat_dict['_'.join(['Total', ws_name_from_method_name])]['count'] += method_count
            
            lst = []
            avg_response_persentil = 0
            cnt = 0
            for l in p['3']['buckets']:
                response_moment_time_ms = int(l['1']['values']['90.0'])/1000
                avg_response_persentil += response_moment_time_ms
                cnt += 1
                lst.append(response_moment_time_ms)

            avg_response_persentil /= cnt

            if avg_response_persentil >= 100:
                avg_response_persentil = round(avg_response_persentil)
            else:
                avg_response_persentil = round(avg_response_persentil, 1)

            wl_stat_dict[method_name]={'count':method_count, 'time':avg_response_persentil}
        

def wl_stat_excel_writer(excel_file_path, wl_stat_dict, file_name_data):
    excel_file = openpyxl.load_workbook(excel_file_path)

    wl_parameter_list = ['count','time']

    for wl_parameter in wl_parameter_list:  

        sheet_name = 'WS_' + wl_parameter
        sheet = excel_file[sheet_name]

        # Подсчет колонок и строк данных на Листе в файле
        rows, cols = sheet.max_row, sheet.max_column

        for col in range(1, cols+1):
            cell_value_in_excel_file = sheet.cell(row=1, column=col).value
            new_cell = sheet.cell(row=rows+1, column=col)
    
            if cell_value_in_excel_file in wl_stat_dict:
                new_cell = sheet.cell(row=rows+1, column=col)
                new_cell.value = wl_stat_dict[cell_value_in_excel_file][wl_parameter]   
                new_cell.alignment = Alignment(
                    horizontal="center", vertical="center")
            try:
                new_cell.value = wl_stat_dict[cell_value_in_excel_file][wl_parameter]
                new_cell.alignment = Alignment(
                    horizontal="center", vertical="center")   
            except: KeyError

            # Запись доп. данных
            if cell_value_in_excel_file in ("Release", "Date", "Test number", "Standart", "Env"):
                cell_value_in_excel_file_format = cell_value_in_excel_file.lower().replace(' ', '_')
                new_cell = sheet.cell(row=rows+1, column=col)
                new_cell.value = file_name_data['CIF'][cell_value_in_excel_file_format]
                new_cell.alignment = Alignment(
                    horizontal="center", vertical="center")   
   
    excel_file.save(excel_file_path)



# test_date = '14.03.2022'
# test_time_start = '14:29'
# test_time_end = '15:29'


# file_name_data = {
#     "CIF": {
#     "env": "L1",
#     "test_number": 5284,
#     "date": '18.01.2022',
#     "release" : "R22.01.2 Final#1"
# },
#     "LAP": {
#     "env": "L1",
#     "standart": 1234,
#     "test_number": 4567,
#     "date": '18.10.2021',
#     'test_time': '15:55-16:55',
#     "url_or_txt": "url",
#     "file_name": "CIF85_AWR_HOMER_4946_vs_4957.html.url",
#     "release" : "21.11.1"
# }
# }



# left_time_bound_prepare = time.strptime(test_date + ' ' + test_time_start, '%d.%m.%Y %H:%M')
# right_time_bound_prepare = time.strptime(test_date + ' ' + test_time_end, '%d.%m.%Y %H:%M')
# struct_left_time_bound = round(time.mktime(left_time_bound_prepare)) * 1000
# struct_right_time_bound = round(time.mktime(right_time_bound_prepare)) * 1000


# index = "load1-homer-*" if file_name_data['CIF']['env'] == 'L1' else "load2-homer-*"


# wl_stat_data_parcer(es, ws_methods_control_list, struct_left_time_bound, struct_right_time_bound, index)

# wl_stat_excel_writer(excel_file_path, wl_stat_dict, file_name_data)

# for key, value in wl_stat_dict.items():
#     print(f"{key}: {value}")

