import re
import openpyxl
from openpyxl.cell import cell
from openpyxl.styles import Alignment



def excel_writer_sms(sms_count_dict: dict, file_name_data: dict, excel_file_path: str) -> None:
    """
    Запись смс в Excel-файл tests_results.xlsx
    Результат работы функции - записанные в файл данные
    """
    excel_file = openpyxl.load_workbook(excel_file_path)
    # Открытие листа SMS в файле Excel
    sheet = excel_file['SMS']

    # Подсчет колонок и строк данных на Листе в фйле
    rows, cols = sheet.max_row, sheet.max_column

    """
    Запись данных по количеству СМС в файл
    Перебор шапки таблицы с названием колонок.
    Если название колонки есть в словаре sms_count_dict, записывает значение по ключу в конец таблицы и удаляет ключ и значение в словаре.
    """

    for col in range(1, cols+1):
        # Запись СМС
        cell_value_in_excel_file = sheet.cell(row=1, column=col).value
        if cell_value_in_excel_file in sms_count_dict:
            new_cell = sheet.cell(row=rows+1, column=col)
            new_cell.value = sms_count_dict.pop(cell_value_in_excel_file)
            new_cell.alignment = Alignment(
                horizontal="center", vertical="center")
        # Запись доп. данных
        if cell_value_in_excel_file in ("Release", "Date", "Test Time", "Test number", "Standart", "Env"):
            cell_value_in_excel_file_format = cell_value_in_excel_file.lower().replace(' ', '_')
            new_cell = sheet.cell(row=rows+1, column=col)
            new_cell.value = file_name_data["LAP"][cell_value_in_excel_file_format]
            new_cell.alignment = Alignment(
                horizontal="center", vertical="center")
                

    excel_file.save(excel_file_path)
    return None


def excel_writer_awr(db_name: str, awr_stat_dict: dict, file_name_data: dict, excel_file_path: str) -> None:
    """
    Запись данных из AWR в Excel-файл tests_results.xlsx
    Результат работы функции - записанные в файл данные
    """
    excel_file = openpyxl.load_workbook(excel_file_path)
    sheet = excel_file[db_name]
    
    # Подсчет колонок и строк данных на Листе в фйле
    rows, cols = sheet.max_row, sheet.max_column

    for col in range(1, cols+1):
        cell_value_in_excel_file = sheet.cell(row=1, column=col).value

            # Запись доп. данных
        if cell_value_in_excel_file in ("Release", "Date", "Test Time", "Test number", "Standart", "Env"):
            cell_value_in_excel_file_format = cell_value_in_excel_file.lower().replace(' ', '_')
            new_cell = sheet.cell(row=rows+1, column=col)
            new_cell.value = file_name_data[db_name][cell_value_in_excel_file_format]
            new_cell.alignment = Alignment(
                horizontal="center", vertical="center")
        
        
        else:
  
            cell_value_in_excel_file_format = cell_value_in_excel_file.lower().replace(' ', '_')
            new_cell = sheet.cell(row=rows+1, column=col)
            new_cell.value = awr_stat_dict['test'][cell_value_in_excel_file_format]
            new_cell.alignment = Alignment(
                horizontal="center", vertical="center")

            # if cell_value_in_excel_file not in ("AWR time", "AWR elapsed time"):
            #     new_cell.number_format = 'Number'


    excel_file.save(excel_file_path)
    return None


node_balance_data_dict = {
    'cif_dbtime_node1_(min)':0,
    'cif_dbtime_node2_(min)':0,
    'cif_avg_active_sessions_node1':0,
    'cif_avg_active_sessions_node2':0,
    'lap_dbtime_node1_(min)':0,
    'lap_dbtime_node2_(min)':0,
    'lap_avg_active_sessions_node1':0,
    'lap_avg_active_sessions_node2':0,

    'awr_time': 0,
    'awr_elapsed_time': 0,
    "release" : 0,
    "test_number": 0,
    "date": 0,
    'test_time': 0,
    "env": 0
}

def node_balance_data_dict_agrigator(db_name: str, awr_stat_dict: dict, node_balance_data_dict: dict, file_name_data: dict):
    if db_name == 'CIF':
        node_balance_data_dict['cif_dbtime_node1_(min)'] = awr_stat_dict['test']['db_time_node1']
        node_balance_data_dict['cif_dbtime_node2_(min)'] = awr_stat_dict['test']['db_time_node2']
        node_balance_data_dict['cif_avg_active_sessions_node1'] = awr_stat_dict['test']['avg_active_sessions_node1']
        node_balance_data_dict['cif_avg_active_sessions_node2'] = awr_stat_dict['test']['avg_active_sessions_node2']

        node_balance_data_dict['awr_time'] = awr_stat_dict['test']['awr_time']
        node_balance_data_dict['awr_elapsed_time'] = awr_stat_dict['test']['awr_elapsed_time']
        node_balance_data_dict['release'] = file_name_data['CIF']['release']
        node_balance_data_dict['test_number'] = file_name_data['CIF']['test_number']
        node_balance_data_dict['date'] = file_name_data['CIF']['date']
        node_balance_data_dict['test_time'] = file_name_data['CIF']['test_time']
        node_balance_data_dict['env'] = file_name_data['CIF']['env']

    elif db_name == 'LAP':
        node_balance_data_dict['lap_dbtime_node1_(min)'] = awr_stat_dict['test']['db_time_node1']
        node_balance_data_dict['lap_dbtime_node2_(min)'] = awr_stat_dict['test']['db_time_node2']
        node_balance_data_dict['lap_avg_active_sessions_node1'] = awr_stat_dict['test']['avg_active_sessions_node1']
        node_balance_data_dict['lap_avg_active_sessions_node2'] = awr_stat_dict['test']['avg_active_sessions_node2']

    return node_balance_data_dict


def data_node_balance_excel_writer(excel_file_path: str, node_balance_data_dict: dict):
    """
    Запись данных из AWR по балансам нагрузки инстансов в Excel-файл tests_results.xlsx
    Результат работы функции - записанные в файл данные
    """
    excel_file = openpyxl.load_workbook(excel_file_path)
    sheet = excel_file['node_balance']
        
    # Подсчет колонок и строк данных на Листе в файле
    rows, cols = sheet.max_row, sheet.max_column

    for col in range(1, cols+1):
        cell_value_in_excel_file = sheet.cell(row=1, column=col).value
        cell_value_in_excel_file_format = cell_value_in_excel_file.lower().replace(' ', '_')
        new_cell = sheet.cell(row=rows+1, column=col)
        try:
            new_cell.value = node_balance_data_dict[cell_value_in_excel_file_format]
            new_cell.alignment = Alignment(
                horizontal="center", vertical="center")   
        except: KeyError

   
    excel_file.save(excel_file_path)
    return None



#excel_writer_awr(db_name, awr_stat_dict, file_name_data)


# excel_writer_sms(sms_count_dict, file_name_data)


# for db_name in file_name_data.keys():
#     excel_writer_sms(data_dict = awr_stat_dict, info_dict = file_name_data[db_name], db_name = db_name)


# def excel_writer1(data_dict: dict, info_dict: dict, db_name: str):
#     print(data_dict)
#     print(info_dict)
#     print(db_name)

#     return None


# for db_name in file_name_data.keys():
#     excel_writer1(data_dict = awr_stat_dict, info_dict = file_name_data[db_name], db_name = db_name)

"""
sms_count_dict = {
    'CREDIT_CARD_RELEASE_IS_APPROVED': 0,
    'CREDIT_IS_APPROVED': 0,
    'CREDIT_IS_ISSUED': 0,
    'CREDIT_IS_ISSUED_ON_OTHER_BANK_ACCOUNT': 0,
    'EHUB_KARTA_POPOLNENIE': 0,
    'EHUB_KARTA_POPOLNENIE_OF': 0,
    'EHUB_KARTA_POPOLNENIE_OFF_PUSH': 0,
    'EHUB_KARTA_POPOLNENIE_PUSH': 0,
    'EHUB_KNPK_POPOLNENIE': 0,
    'EHUB_KNPK_POPOLNENIE_OF': 0,
    'EHUB_KNPK_POPOLNENIE_OF_PUSH': 0,
    'EHUB_KNPK_POPOLNENIE_PUSH': 0,
    'HOMER_CREDITCARD_REJECT': 0,
    'PK_NK_APPR': 0,
    'PK_NK_PRE_APPR': 0,
    'PK_NK_WLCM': 0,
    'RK_APPR': 0,
    'RK_WLCM': 0
}

file_name_data = {
    "CIF": {
        "env": "L1",
        "standart": 1234,
        "test_number": 4567,
        "date": '18.10.2021',
        'test_time': '15:55-16:55',
        "url_or_txt": "url",
        "file_name": "CIF85_AWR_HOMER_4946_vs_4957.html.url",
        "release" : "21.11.1"
    },
    "LAP": {
        "env": "L1",
        "standart": 1234,
        "test_number": 4567,
        "date": '18.10.2021',
        'test_time': '15:55-16:55',
        "url_or_txt": "url",
        "file_name": "LAP85_AWR_HOMER_4946_vs_4957.html.url",
        "release" : "21.11.1"
    }
}

awr_stat_dict = {
    'standart': {
        'db_time': 300,
        'cpu_time': 0,
        'elapsed_time': 0,
        'io_time': 0,
        'buffer_gets': 0,
        'physical_reads': 0,
        'captured_sql_executions': 0,
        'cluster_wait_time': 0,
        'awr_elapsed_time': 0,
        'awr_time': 0,
        'db_time_node1': 190,
        'db_time_node2': 110,
        'avg_active_sessions_node1':1,
        'avg_active_sessions_node2':2
    },
    'test': {
        'db_time': 299,
        'cpu_time': 0,
        'elapsed_time': 0,
        'io_time': 0,
        'buffer_gets': 0,
        'physical_reads': 0,
        'captured_sql_executions': 0,
        'cluster_wait_time': 0,
        'awr_elapsed_time': 0,
        'awr_time': 0,
        'db_time_node1': 200,
        'db_time_node2': 99,
        'avg_active_sessions_node1':3,
        'avg_active_sessions_node2':11
    }
}

db_name = "CIF"

date = '16.10.2021'
time = '12:54-13:54'
test_number = '7777'
standart = '9999'
env = 'L1'
excel_file_path = 'data/new/tests_results.xlsx'

file_name_data = {
    "CIF": {
        "env": "L1",
        "standart": 1234,
        "test_number": 4567,
        "date": '18.10.2021',
        'test_time': '15:55-16:55',
        "url_or_txt": "url",
        "file_name": "CIF85_AWR_HOMER_4946_vs_4957.html.url"
    },
    "LAP": {
        "env": "L1",
        "standart": 1234,
        "test_number": 4567,
        "date": '18.10.2021',
        'test_time': '15:55-16:55',
        "url_or_txt": "url",
        "file_name": "LAP85_AWR_HOMER_4946_vs_4957.html.url"
    }
}

"""